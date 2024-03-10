import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Adjust file paths for the Jupyter environment
csv_file_path = '/Users/darren/Desktop/gantt/oricsv.csv'
xlsx_file_path = '/Users/darren/Desktop/gantt/xlsFromCsv.xlsx'
task_name = 'Task Name'
date_start = 'Date Start'
date_end = 'Date End'
task_id = 'Task Id'  # Assuming this is the name of your Task ID column

# Load the CSV file
df = pd.read_csv(csv_file_path, parse_dates=[date_start, date_end])
tasks = df[task_id].unique()
date_range = pd.date_range(df[date_start].min(), df[date_end].max())

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Define a font with size 8 and light blue fill
font_size_8 = Font(size=8)
light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

# Set column headers with font size 8
ws.append([task_id, task_name, date_start, date_end] + [d.strftime('%Y-%m-%d') for d in date_range])
for cell in ws[1]:
    cell.font = font_size_8
    cell.alignment = Alignment(horizontal='center')

# Map each task to a row number
task_to_row = {task_id: idx + 2 for idx, task_id in enumerate(df[task_id].unique(), start=1)}

# Populate the task name, start and end dates, and the task ID
for task_id in df[task_id].unique():
    task_row = task_to_row[task_id]
    task_data = df[df[task_id] == task_id]
    ws.cell(row=task_row, column=1, value=task_id)
    ws.cell(row=task_row, column=2, value=task_data[task_name].iloc[0])
    ws.cell(row=task_row, column=3, value=task_data[date_start].dt.date.iloc[0])
    ws.cell(row=task_row, column=4, value=task_data[date_end].dt.date.iloc[0])

# Apply the blue fill for the duration of each task
for _, row in df.iterrows():
    task = row[task_name]
    start = row[date_start].to_pydatetime().date()
    end = row[date_end].to_pydatetime().date()
    task_row = task_to_row[task]
    for col, date in enumerate(date_range, start=5):  # Start from 5 to account for the new Task ID column
        cell = ws.cell(row=task_row, column=col)
        cell.font = font_size_8  # Set font size for all cells
        if start <= date.date() <= end:
            cell.fill = light_blue_fill  # Apply fill for the task's duration

# Save the workbook as an XLSX file
wb.save(xlsx_file_path)
print("File saved at:", xlsx_file_path)

